import sys
import csv
import openpyxl
import os

def convert_xlsx_to_csv(xlsx_file, csv_path):
    wb = openpyxl.load_workbook(xlsx_file, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    with open(csv_path, "w", newline='', encoding="utf-8") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
    print(f"Converted {xlsx_file} to {csv_path}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python3 convert_xlsx_to_csv.py input.xlsx output.csv")
        sys.exit(1)

    xlsx_file = sys.argv[1]
    csv_path = sys.argv[2]

    convert_xlsx_to_csv(xlsx_file, csv_path)